"""Excel export builders for the reports slice (D3).

The fridge-report workbook is built in two passes, exactly as required:

* **Polars** ``DataFrame.write_excel`` renders the per-product table body as a
  proper Excel table, anchored a few rows down to leave the top of the sheet
  free.
* **openpyxl** re-opens that workbook and writes the summary block (title,
  fridge, period and the four KPI figures) into the reserved top rows, so the
  summary reads first and the table follows below it.

Keeping the two engines in separate passes (Polars writes, openpyxl decorates)
avoids mixing incompatible workbook objects and yields a plain ``.xlsx`` any
consumer — including openpyxl itself — can open.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal

import polars as pl
from openpyxl import load_workbook
from openpyxl.styles import Font

from app.money import cents_to_euro_decimal
from app.services.finance_service import FridgeReportData

_SHEET_NAME = "Fridge Report"
# The table body starts here; rows 1..10 are reserved for the summary block.
_TABLE_ANCHOR = "A11"
_EURO_FORMAT = "#,##0.00"
_QTY_FORMAT = "#,##0"

_TABLE_COLUMNS = (
    "Product",
    "Code",
    "Category",
    "Added Qty",
    "Unit Buying Price (EUR)",
    "Unit Selling Price (EUR)",
)


@dataclass(frozen=True)
class ExcelDocument:
    """A rendered workbook plus its download filename."""

    filename: str
    content: bytes

    @property
    def media_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _euro(cents: int) -> float:
    """Cents → euros as a float for a numeric Excel cell."""
    return float(cents_to_euro_decimal(cents))


def _build_table_frame(data: FridgeReportData) -> pl.DataFrame:
    return pl.DataFrame(
        {
            "Product": [row.name for row in data.rows],
            "Code": [row.code for row in data.rows],
            "Category": [row.category or "" for row in data.rows],
            "Added Qty": [row.added_qty for row in data.rows],
            "Unit Buying Price (EUR)": [_euro(row.unit_buying_price) for row in data.rows],
            "Unit Selling Price (EUR)": [
                _euro(row.unit_selling_price) for row in data.rows
            ],
        },
        schema={
            "Product": pl.Utf8,
            "Code": pl.Utf8,
            "Category": pl.Utf8,
            "Added Qty": pl.Int64,
            "Unit Buying Price (EUR)": pl.Float64,
            "Unit Selling Price (EUR)": pl.Float64,
        },
    )


def _write_summary_block(worksheet, data: FridgeReportData) -> None:
    """Fill the reserved top rows with the title + fridge + period + KPIs."""
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    worksheet["A1"] = "Fridge Report"
    worksheet["A1"].font = title_font

    worksheet["A2"] = "Fridge"
    worksheet["A2"].font = bold
    worksheet["B2"] = f"{data.fridge_name} (#{data.fridge_id})"

    worksheet["A3"] = "Period"
    worksheet["A3"].font = bold
    worksheet["B3"] = f"{data.date_from.isoformat()} → {data.date_to.isoformat()}"

    margin_pct = (
        f"{data.margin_pct}%" if data.margin_pct is not None else "n/a"
    )
    kpis: list[tuple[str, object, str | None]] = [
        ("Total Added Qty", data.total_added_qty, _QTY_FORMAT),
        ("Food Cost (EUR)", _euro(data.food_cost), _EURO_FORMAT),
        ("Revenue (EUR)", _euro(data.revenue), _EURO_FORMAT),
        ("Food Margin (EUR)", _euro(data.margin), _EURO_FORMAT),
        ("Food Margin %", margin_pct, None),
    ]
    for index, (label, value, number_format) in enumerate(kpis):
        row_no = 5 + index
        label_cell = worksheet.cell(row=row_no, column=1, value=label)
        label_cell.font = bold
        value_cell = worksheet.cell(row=row_no, column=2, value=value)
        if number_format is not None:
            value_cell.number_format = number_format


def build_fridge_report_xlsx(data: FridgeReportData) -> ExcelDocument:
    """Render the fridge report to a streamable ``.xlsx`` document."""
    frame = _build_table_frame(data)

    buffer = io.BytesIO()
    frame.write_excel(
        buffer,
        worksheet=_SHEET_NAME,
        position=_TABLE_ANCHOR,
        table_style="Table Style Medium 9",
        autofit=True,
        column_formats={
            "Added Qty": _QTY_FORMAT,
            "Unit Buying Price (EUR)": _EURO_FORMAT,
            "Unit Selling Price (EUR)": _EURO_FORMAT,
        },
    )

    buffer.seek(0)
    workbook = load_workbook(buffer)
    worksheet = workbook[_SHEET_NAME]
    _write_summary_block(worksheet, data)

    out = io.BytesIO()
    workbook.save(out)

    filename = (
        f"fridge-report_{data.fridge_id}_"
        f"{data.date_from.isoformat()}_{data.date_to.isoformat()}.xlsx"
    )
    return ExcelDocument(filename=filename, content=out.getvalue())
